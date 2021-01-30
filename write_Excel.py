import os.path
import tkinter
import xlsxwriter
import openpyxl
from tkinter import *
from tkinter.filedialog import askdirectory

from datetime import  datetime
class write_Data():
    file_path = ""
    def set_location(self,path):

        self.file_path = path

        return  self.file_path

    def write_xlsx(self,array, platform):

        xlsx_file_name = platform + "_data.xlsx"
        if "/" in self.file_path:
            file_name_path = self.file_path + "/" + xlsx_file_name
        elif "\\" in self.file_path:
            file_name_path = self.file_path + "\\" + xlsx_file_name

        sheet_name = "comparision group"
        row = 0
        a_row, a_col = array.shape
        if(not os.path.isfile(file_name_path) ):
            workbook = xlsxwriter.Workbook(file_name_path)
            ws = workbook.add_worksheet(name= sheet_name)
            workbook.close()

        xfile = openpyxl.load_workbook(file_name_path)
        if sheet_name in xfile.sheetnames:
            sheet = xfile.get_sheet_by_name(sheet_name)
        else:
            sheet = xfile.create_sheet(sheet_name)
        row = sheet.max_row

        sheet.cell(row=row,column=1, value=str(datetime.now().strftime("%d-%b-%Y (%H:%M)")))

        for x in range(2022-2008):
            sheet.cell(row=row, column=x+2, value=(2008+x) )
        sheet.cell(row=row + 1, column=1, value="Purchase Pending")
        sheet.cell(row=row + 2, column=1, value="Incoming Cars")
        sheet.cell(row=row + 3, column=1, value="Available")
        sheet.cell(row=row + 4, column=1, value="Sold")
        sheet.cell(row=row + 5, column=1, value="Available to pending")
        sheet.cell(row=row + 6, column=1, value="Pending to available")
        sheet.cell(row=row + 7, column=1, value="Incoming to availble")
        sheet.cell(row=row + 8, column=1, value="Pending to sold")

        for x in range (a_row ):
            for y in range( a_col):
                sheet.cell(row=row +1+x , column=y + 2, value=array[x][y])

        xfile.save(filename=file_name_path)
    def write_txt(self,list, platform):
        txt_file_name = platform + "_data.txt"
        if "/" in self.file_path:
            file_name_path = self.file_path + "/" + txt_file_name
        elif "\\" in self.file_path:
            file_name_path = self.file_path + "\\" + txt_file_name
        file = open(file_name_path,"w+")
        for car in list:
            file.write(car.toString())
        print("closing")
        file.close()